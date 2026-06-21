"""
Export Excel peta referensi 3 sumber per fitur (MODUL 9).
Menghasilkan .xlsx dari REF_DATA agar klien dapat memeriksa & memberi feedback.
REF_DATA dijaga sinkron dengan static/js/pages/referensi_page.js.
"""

import io

from flask import send_file

from utils.response import error


# Peta referensi per fitur: (fitur, rancangan/spreadsheet, buku, pikobe/ledik).
REF_DATA = [
    ("Identitas Prodi", "\u2014", "Tabel A Isian identitas (hal 6)", "\u2014"),
    ("Profil Lulusan", "\u2014", "Tabel 1 Profil Lulusan (hal 15)", "\u2014"),
    ("CPL Prodi", "Sheet 3. CPL Prodi", "Tabel 2 CPL Kompetensi Utama (hal 17)", "Tabel 2 Capaian Pembelajaran Lulusan"),
    ("Bahan Kajian", "Sheet 6. Bahan Kajian", "Tabel 4 Rumusan Bahan Kajian (hal 19-20)", "\u2014"),
    ("Mata Kuliah", "Sheet 11. Susunan Mata Kuliah", "Tabel 9 Susunan Mata Kuliah (hal 26-27)", "\u2014"),
    ("CPMK", "\u2014", "Tabel 12 Pemetaan CPL-CPMK-MK (hal 31)", "\u2014"),
    ("Matriks CPL - PL", "\u2014", "Tabel 3 Pemetaan CPL dan PL (hal 18)", "\u2014"),
    ("Matriks CPL - BK", "\u2014", "Tabel 5 Pemetaan CPL-BK (hal 21)", "\u2014"),
    ("Matriks BK - MK", "\u2014", "Tabel 6 Pemetaan BK-MK (hal 22)", "\u2014"),
    ("Matriks CPL - MK", "\u2014", "Tabel 7 Pemetaan CPL-MK (hal 24)", "\u2014"),
    ("Matriks CPMK - MK", "\u2014", "Tabel 14 Pemetaan MK-CPL-CPMK (hal 33)", "\u2014"),
    ("Pemetaan CPL-BK-MK", "\u2014", "Tabel 8 Pemetaan BK-CPL-MK (hal 24)", "\u2014"),
    ("Organisasi MK", "\u2014", "Tabel 10 Organisasi Mata Kuliah (hal 29)", "\u2014"),
    ("Peta Pemenuhan CPL", "\u2014", "Tabel 11 Peta Pemenuhan CPL (hal 30)", "\u2014"),
    ("MK - CPMK - Sub CPMK", "\u2014", "Tabel 15 Pemetaan MK-CPMK-Sub-CPMK (hal 35)", "\u2014"),
    ("RPS", "Sheet 21. Rancangan RPS", "4. RPS (hal 42-45)", "Sheet Contoh RPS"),
]

HEADERS = ["Sidebar / Fitur", "Rancangan Kurikulum", "Buku Kurikulum", "PIKOBE / LEDIK"]


def export_referensi():
    """GET /api/referensi/export -- unduh peta referensi sebagai .xlsx."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return error("Fitur export membutuhkan paket 'openpyxl' yang belum terpasang.", status=501)

    wb = Workbook()
    ws = wb.active
    ws.title = "Peta Referensi"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3A5F")
    wrap = Alignment(vertical="top", wrap_text=True)

    for col_idx, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="center")

    for row_idx, row in enumerate(REF_DATA, start=2):
        for col_idx, value in enumerate(row, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.alignment = wrap

    widths = [26, 30, 42, 38]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="peta_referensi_kurikulum_obe.xlsx",
    )


ROUTE_DEFINITIONS = [
    ("GET", "/api/referensi/export", export_referensi, "view_kurikulum"),
]
