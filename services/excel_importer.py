"""
Parser Excel untuk import data awal dari file Rancangan Kurikulum.
"""

import state


def import_from_excel(file_path):
    """
    Import data dari file Excel.
    Membutuhkan openpyxl.

    Sheet yang diproses:
    - Sheet 1: Identitas Prodi
    - Sheet 2-5: CPL, BK
    - Sheet 6-10: MK, Matriks
    - Sheet 11-18: CPMK, Penilaian, Bobot
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": "openpyxl belum terinstall"}

    wb = load_workbook(file_path, data_only=True)
    results = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cleaned = [_clean_cell(c) for c in row]
            if any(c for c in cleaned):
                rows.append(cleaned)
        results[sheet_name] = rows

    wb.close()
    return results


def _clean_cell(value):
    """Membersihkan nilai sel."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value
